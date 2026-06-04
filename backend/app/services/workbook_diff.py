"""Pure diff / classify / marshal engine for live workbook updates.

This module is the brain behind `files_smart_update`. It is deliberately
**Graph-free and asyncio-free**: it takes two `.xlsx` byte blobs (the freshly
built *proposed* workbook and the *live* workbook downloaded from SharePoint)
plus an optional per-tab `region_map`, and returns an :class:`UpdatePlan`
describing what — if anything — can be safely applied to the open workbook via
Graph Excel *session* operations.

Design (see ~/marvin/docs/live-workbook-update-spec.md and the approved plan):

Governing principle — a change is **live-editable** iff the proposed workbook
can be reached from the live one by Graph session operations we can both
reliably *detect* from the bytes and faithfully *replay*: value/formula writes,
content clears, and worksheet add/delete/rename/reorder. Graph's blind spots
(cell styles, number formats, conditional formatting, merges, column widths,
freeze/filter, charts/images/pivots/validation) cannot be reproduced from
openpyxl bytes, so a change that *requires* reproducing them defers.

Formatting is treated as a **boundary we never cross**, not a "changed -> defer"
trigger (the pragmatic stance): we leave the live file's formatting / conditional
formatting exactly as-is and only ever write values, formulas and structure. A
formatting difference only forces a defer when there is *nothing else* to apply
(the "format-only run -> defer + signal" case), so the caller can fall back to a
clean closed-file replace.

Everything here is unit-testable from two `BytesIO` blobs with no Graph mock.
"""

from __future__ import annotations

import math
from dataclasses import dataclass, field
from datetime import date, datetime, time
from io import BytesIO
from typing import Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.utils.datetime import to_excel

# Array / data-table formulas can't be reproduced via a single-cell range write.
try:  # openpyxl >= 3.1
    from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula
    _ARRAY_FORMULA_TYPES: tuple = (ArrayFormula, DataTableFormula)
except Exception:  # pragma: no cover - defensive for older openpyxl
    _ARRAY_FORMULA_TYPES = ()


# Soft guard: refuse to brute-force-diff a pathologically large grid.
_MAX_DIFF_CELLS = 2_000_000
# A region whose payload would be huge is split into horizontal stripes.
_STRIPE_ROWS = 500
# Rename inference: how similar two odd-one-out sheets must be to be a rename,
# and how much the best candidate must beat the runner-up to be unambiguous.
_RENAME_MATCH_THRESHOLD = 0.6
_RENAME_MARGIN = 0.15


# ---------------------------------------------------------------------------
# Plan dataclasses
# ---------------------------------------------------------------------------


@dataclass
class Box:
    """A 1-based inclusive cell rectangle, as parsed from an A1 range."""

    min_row: int
    min_col: int
    max_row: int
    max_col: int

    def contains(self, row: int, col: int) -> bool:
        return (
            self.min_row <= row <= self.max_row
            and self.min_col <= col <= self.max_col
        )

    def address(self) -> str:
        return (
            f"{get_column_letter(self.min_col)}{self.min_row}:"
            f"{get_column_letter(self.max_col)}{self.max_row}"
        )


@dataclass
class RangeWrite:
    """Write a 2-D `values` array into `sheet`!`address` (values pre-marshalled)."""

    sheet: str
    address: str
    values: list


@dataclass
class RenameOp:
    old: str
    new: str


@dataclass
class AddOp:
    """Add `name`, then write its content (range writes already marshalled)."""

    name: str
    writes: list = field(default_factory=list)


@dataclass
class DeleteOp:
    name: str


@dataclass
class ReorderOp:
    name: str
    position: int  # 0-based target position


@dataclass
class UpdatePlan:
    mode: str  # "LIVE" or "DEFER"
    reason: str
    renames: list = field(default_factory=list)
    adds: list = field(default_factory=list)
    deletes: list = field(default_factory=list)
    reorders: list = field(default_factory=list)
    range_writes: list = field(default_factory=list)

    @property
    def ranges_written(self) -> int:
        n = len(self.range_writes)
        for add in self.adds:
            n += len(add.writes)
        return n

    @property
    def has_ops(self) -> bool:
        return bool(
            self.renames
            or self.adds
            or self.deletes
            or self.reorders
            or self.range_writes
        )


class _Defer(Exception):
    """Internal short-circuit: abandon classification and defer with a reason."""

    def __init__(self, reason: str):
        super().__init__(reason)
        self.reason = reason


# ---------------------------------------------------------------------------
# Region map
# ---------------------------------------------------------------------------


def parse_region_map(region_map: Optional[dict]) -> dict:
    """Turn {sheet: {"data": "A2:U200"}} (or {sheet: [ranges]}) into {sheet: [Box]}.

    Accepts either a single ``data`` range, a list of ranges, or a list of
    ``{"data": range}`` dicts per sheet, for forward-compatibility.
    """
    boxes: dict = {}
    if not region_map:
        return boxes
    for sheet, spec in region_map.items():
        ranges: list = []
        if isinstance(spec, str):
            ranges = [spec]
        elif isinstance(spec, dict):
            data = spec.get("data")
            if isinstance(data, str):
                ranges = [data]
            elif isinstance(data, (list, tuple)):
                ranges = list(data)
        elif isinstance(spec, (list, tuple)):
            for entry in spec:
                if isinstance(entry, str):
                    ranges.append(entry)
                elif isinstance(entry, dict) and isinstance(entry.get("data"), str):
                    ranges.append(entry["data"])
        sheet_boxes = []
        for addr in ranges:
            min_col, min_row, max_col, max_row = range_boundaries(addr)
            sheet_boxes.append(Box(min_row, min_col, max_row, max_col))
        if sheet_boxes:
            boxes[sheet] = sheet_boxes
    return boxes


# ---------------------------------------------------------------------------
# Value normalisation (diff) and marshalling (write)
# ---------------------------------------------------------------------------


def _is_formula(value) -> bool:
    if isinstance(value, _ARRAY_FORMULA_TYPES):
        return True
    return isinstance(value, str) and value.startswith("=")


def normalize_value(value):
    """Canonicalise a cell value for *comparison* so we don't report phantom diffs.

    Collapses ``""`` <-> ``None``, ``int`` <-> ``float``, and keys by type so a
    string ``"5"`` never equals a number ``5``. Strings are NOT stripped —
    trailing spaces can be meaningful.
    """
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return ("b", value)
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return ("x", "nan")
        return ("n", float(value))
    if isinstance(value, (datetime, date, time)):
        return ("d", value.isoformat() if hasattr(value, "isoformat") else str(value))
    if isinstance(value, _ARRAY_FORMULA_TYPES):
        # Reference text uniquely identifies the array formula for diff purposes.
        return ("af", getattr(value, "text", str(value)))
    if isinstance(value, str):
        return ("s", value)
    return ("o", str(value))


def marshal_value(value, epoch):
    """Convert an openpyxl cell value into a JSON scalar Graph accepts.

    Returns number | str | bool | None. Raises :class:`_Defer` for values that
    cannot be reproduced via a range write (array/data-table formulas).
    """
    if value is None or value == "":
        return None
    if isinstance(value, bool):  # before int — bool is a subclass of int
        return value
    if isinstance(value, (datetime, date, time)):
        dt = value
        if isinstance(dt, datetime) and dt.tzinfo is not None:
            dt = dt.replace(tzinfo=None)
        return to_excel(dt, epoch)
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return None
        return value
    if isinstance(value, _ARRAY_FORMULA_TYPES):
        raise _Defer("proposed workbook uses an array/data-table formula")
    if isinstance(value, str):
        return value  # formula ("=...") or text; Graph treats leading "=" as a formula
    return str(value)


# ---------------------------------------------------------------------------
# Worksheet structure helpers (only the RELIABLE signals)
# ---------------------------------------------------------------------------


def _merged_set(ws) -> set:
    return {str(r) for r in ws.merged_cells.ranges}


def _freeze(ws):
    fp = ws.freeze_panes
    return None if fp in (None, "A1") else fp


def _autofilter(ws):
    return ws.auto_filter.ref


def _has_unreproducible_structure(ws) -> bool:
    """True if the sheet carries structure a value/formula write can't reproduce.

    Used to gate a *newly added* sheet (we can only create + write values), and
    to flag a format-only change on an existing sheet. Restricted to reliable
    signals — no font/fill/border comparison (those false-positive between a
    freshly built workbook and a SharePoint-saved one).
    """
    if _merged_set(ws):
        return True
    if _freeze(ws) is not None:
        return True
    if _autofilter(ws) is not None:
        return True
    try:
        if list(ws.conditional_formatting):
            return True
    except Exception:
        pass
    if getattr(ws, "_charts", None):
        return True
    if getattr(ws, "_images", None):
        return True
    dvs = getattr(getattr(ws, "data_validations", None), "dataValidation", None)
    if dvs:
        return True
    for cd in ws.column_dimensions.values():
        if cd.width is not None and cd.customWidth:
            return True
    return False


def _reliable_structure_differs(pws, lws) -> bool:
    """Compare only the trustworthy structure/format signals between two sheets."""
    if _merged_set(pws) != _merged_set(lws):
        return True
    if _freeze(pws) != _freeze(lws):
        return True
    if _autofilter(pws) != _autofilter(lws):
        return True
    if _number_format_grid_differs(pws, lws):
        return True
    return False


def _number_format_grid_differs(pws, lws) -> bool:
    max_row = max(pws.max_row, lws.max_row)
    max_col = max(pws.max_column, lws.max_column)
    if max_row * max_col > _MAX_DIFF_CELLS:
        return False  # too large to scan; rely on value diff only
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            if _norm_fmt(pws.cell(r, c).number_format) != _norm_fmt(
                lws.cell(r, c).number_format
            ):
                return True
    return False


def _norm_fmt(fmt) -> str:
    if fmt in (None, "", "General"):
        return "General"
    return fmt


# ---------------------------------------------------------------------------
# Cell diff
# ---------------------------------------------------------------------------


def _cell_kind(pc, lc):
    """Return None | 'VALUE' | 'FORMULA' for a (proposed, live) cell pair.

    Raises :class:`_Defer` if the proposed cell is an array/data-table formula.
    """
    pv, lv = pc.value, lc.value
    if isinstance(pv, _ARRAY_FORMULA_TYPES):
        raise _Defer("proposed workbook uses an array/data-table formula")
    p_form = _is_formula(pv)
    l_form = _is_formula(lv)
    if normalize_value(pv) == normalize_value(lv) and p_form == l_form:
        return None
    return "FORMULA" if (p_form or l_form) else "VALUE"


def _diff_sheet_cells(pws, lws, boxes, sheet_name):
    """Diff one matched sheet pair. Returns the list of in-region Box objects
    that contain at least one change. Defers on any out-of-region change."""
    max_row = max(pws.max_row, lws.max_row)
    max_col = max(pws.max_column, lws.max_column)
    if max_row * max_col > _MAX_DIFF_CELLS:
        raise _Defer(f"'{sheet_name}' is too large to diff safely")

    touched_boxes: list = []
    touched_idx: set = set()
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            kind = _cell_kind(pws.cell(r, c), lws.cell(r, c))
            if kind is None:
                continue
            box_i = _find_box(boxes, r, c)
            if box_i is None:
                raise _Defer(
                    f"change outside declared region at "
                    f"{sheet_name}!{get_column_letter(c)}{r}"
                )
            if box_i not in touched_idx:
                touched_idx.add(box_i)
                touched_boxes.append(boxes[box_i])
    return touched_boxes


def _find_box(boxes, row, col) -> Optional[int]:
    for i, b in enumerate(boxes):
        if b.contains(row, col):
            return i
    return None


def _box_writes(ws, box, sheet_name, epoch) -> list:
    """Build RangeWrite(s) for a region box: the proposed values for the whole
    box (changed or not — rewriting identical cells is a no-op; `None` clears).
    Row-band split if the box is large."""
    writes = []
    rows = box.max_row - box.min_row + 1
    cols = box.max_col - box.min_col + 1
    stripe = _STRIPE_ROWS if rows * cols > _STRIPE_ROWS * 64 else rows
    start = box.min_row
    while start <= box.max_row:
        end = min(start + stripe - 1, box.max_row)
        values = []
        for r in range(start, end + 1):
            row_vals = [
                marshal_value(ws.cell(r, c).value, epoch)
                for c in range(box.min_col, box.max_col + 1)
            ]
            values.append(row_vals)
        addr = (
            f"{get_column_letter(box.min_col)}{start}:"
            f"{get_column_letter(box.max_col)}{end}"
        )
        writes.append(RangeWrite(sheet_name, addr, values))
        start = end + 1
    return writes


# ---------------------------------------------------------------------------
# Worksheet reconciliation (add / delete / rename / reorder)
# ---------------------------------------------------------------------------


def _sheet_signature(ws) -> set:
    """A position-insensitive multiset-ish signature of a sheet's non-empty
    values, used to infer renames robustly against row shifts."""
    sig = set()
    max_row = min(ws.max_row, 400)
    max_col = min(ws.max_column, 64)
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            nv = normalize_value(ws.cell(r, c).value)
            if nv is not None:
                sig.add(nv)
    return sig


def _similarity(a: set, b: set) -> float:
    if not a and not b:
        return 1.0
    if not a or not b:
        return 0.0
    inter = len(a & b)
    union = len(a | b)
    return inter / union if union else 0.0


def _reconcile_sheets(prop, live):
    """Match worksheets across the two workbooks.

    Returns (pairs, renames, adds, deletes) where pairs is a list of
    (prop_name, live_name) for sheets to cell-diff. Raises _Defer on an
    ambiguous rename or a styled new tab we can't reproduce.
    """
    prop_names = prop.sheetnames
    live_names = live.sheetnames
    common = [n for n in prop_names if n in live_names]
    prop_only = [n for n in prop_names if n not in live_names]
    live_only = [n for n in live_names if n not in prop_names]

    pairs = [(n, n) for n in common]
    renames: list = []

    # Infer renames by content similarity among the odd-one-out sheets.
    if prop_only and live_only:
        psig = {n: _sheet_signature(prop[n]) for n in prop_only}
        lsig = {n: _sheet_signature(live[n]) for n in live_only}
        remaining_live = set(live_only)
        # Greedy: strongest matches first.
        candidates = []
        for pn in prop_only:
            for ln in live_only:
                candidates.append((_similarity(psig[pn], lsig[ln]), pn, ln))
        candidates.sort(reverse=True, key=lambda t: t[0])
        matched_prop: dict = {}
        for score, pn, ln in candidates:
            if pn in matched_prop or ln not in remaining_live:
                continue
            if score < _RENAME_MATCH_THRESHOLD:
                continue
            # Ambiguity guard: is there a near-tie for this prop sheet?
            rivals = [
                s for s, p2, l2 in candidates
                if p2 == pn and l2 != ln and l2 in remaining_live
            ]
            if rivals and (score - max(rivals)) < _RENAME_MARGIN:
                raise _Defer(f"ambiguous sheet rename for '{pn}'")
            matched_prop[pn] = ln
            remaining_live.discard(ln)
            renames.append(RenameOp(old=ln, new=pn))
            pairs.append((pn, ln))
        prop_only = [n for n in prop_only if n not in matched_prop]
        live_only = list(remaining_live)

    adds = []
    for n in prop_only:
        if _has_unreproducible_structure(prop[n]):
            raise _Defer(f"new sheet '{n}' has formatting/objects we can't reproduce")
        adds.append(n)
    deletes = [DeleteOp(name=n) for n in live_only]
    return pairs, renames, adds, deletes


def _added_sheet_writes(ws, name, epoch) -> list:
    """Write the entire used range of a newly added (plain) sheet."""
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row == 1 and max_col == 1 and ws.cell(1, 1).value in (None, ""):
        return []  # empty sheet
    box = Box(1, 1, max_row, max_col)
    return _box_writes(ws, box, name, epoch)


def _reorders(prop, live, renames, deletes) -> list:
    """Emit ReorderOps if the proposed order of the surviving sheets differs."""
    rename_map = {r.old: r.new for r in renames}  # live name -> prop name
    deleted = {d.name for d in deletes}
    # Live sheets that survive, mapped to their proposed names, in LIVE order.
    live_order = [
        rename_map.get(n, n) for n in live.sheetnames if n not in deleted
    ]
    # Their target order = proposed order restricted to those names.
    survivors = set(live_order)
    target_order = [n for n in prop.sheetnames if n in survivors]
    if live_order == target_order:
        return []
    return [ReorderOp(name=n, position=i) for i, n in enumerate(target_order)]


# ---------------------------------------------------------------------------
# Top-level classify
# ---------------------------------------------------------------------------


def classify(new_bytes: bytes, live_bytes: bytes, region_map: Optional[dict]) -> UpdatePlan:
    """Diff proposed vs live and return an :class:`UpdatePlan`.

    `mode == "LIVE"` with ops -> apply them surgically. `mode == "LIVE"` with no
    ops -> already up to date. `mode == "DEFER"` -> caller should fall back to a
    closed-file replace (and notify).
    """
    try:
        prop = load_workbook(BytesIO(new_bytes), data_only=False)
        live = load_workbook(BytesIO(live_bytes), data_only=False)
        boxes_by_sheet = parse_region_map(region_map)
        epoch = prop.epoch

        try:
            pairs, renames, add_names, deletes = _reconcile_sheets(prop, live)
        except _Defer as d:
            return UpdatePlan("DEFER", d.reason)

        range_writes: list = []
        for prop_name, live_name in pairs:
            boxes = boxes_by_sheet.get(prop_name, [])
            touched = _diff_sheet_cells(
                prop[prop_name], live[live_name], boxes, prop_name
            )
            for box in touched:
                range_writes.extend(
                    _box_writes(prop[prop_name], box, prop_name, epoch)
                )

        adds = []
        for name in add_names:
            adds.append(AddOp(name=name, writes=_added_sheet_writes(prop[name], name, epoch)))

        reorders = _reorders(prop, live, renames, deletes)

        plan = UpdatePlan(
            "LIVE",
            "value/formula edits in-region",
            renames=renames,
            adds=adds,
            deletes=deletes,
            reorders=reorders,
            range_writes=range_writes,
        )

        if plan.has_ops:
            return plan

        # Nothing applicable. Is there a non-reproducible (format/structure)
        # difference? If so, signal a defer so a clean replace can run; else the
        # open workbook already matches the proposed build.
        for prop_name, live_name in pairs:
            if _reliable_structure_differs(prop[prop_name], live[live_name]):
                return UpdatePlan("DEFER", "formatting/structural change only")
        return UpdatePlan("LIVE", "already up to date")
    except _Defer as d:  # raised mid-marshal (e.g. array formula)
        return UpdatePlan("DEFER", d.reason)
