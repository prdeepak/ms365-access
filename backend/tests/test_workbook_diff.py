"""Unit tests for the pure diff/classify/marshal engine (no Graph)."""

import math
from datetime import datetime
from io import BytesIO

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.formula import ArrayFormula

from app.services import workbook_diff as wd


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def wb_bytes(sheets, *, number_formats=None, styler=None):
    """Build an .xlsx blob. `sheets` = {name: [[row values], ...]}.

    `number_formats` = {name: {(row, col): "fmt"}}. `styler` = fn(wb) for
    structural tweaks (merges, freeze, charts, etc.).
    """
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for r, row in enumerate(rows, start=1):
            for c, val in enumerate(row, start=1):
                if val is not None:
                    ws.cell(row=r, column=c, value=val)
        if number_formats and name in number_formats:
            for (r, c), fmt in number_formats[name].items():
                ws.cell(row=r, column=c).number_format = fmt
    if styler:
        styler(wb)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def flat_writes(plan):
    writes = list(plan.range_writes)
    for add in plan.adds:
        writes.extend(add.writes)
    return writes


# ---------------------------------------------------------------------------
# Cell value / formula diffs
# ---------------------------------------------------------------------------


def test_value_change_in_region_is_live():
    live = wb_bytes({"S": [[1, 10], [2, 20]]})
    prop = wb_bytes({"S": [[1, 10], [2, 99]]})
    plan = wd.classify(prop, live, {"S": {"data": "A1:B2"}})
    assert plan.mode == "LIVE"
    assert len(plan.range_writes) == 1
    w = plan.range_writes[0]
    assert w.sheet == "S"
    assert w.address == "A1:B2"
    assert w.values == [[1, 10], [2, 99]]
    assert plan.ranges_written == 1


def test_formula_change_in_region_is_live():
    live = wb_bytes({"S": [["x", "=A1"]]})
    prop = wb_bytes({"S": [["x", "=A1+1"]]})
    plan = wd.classify(prop, live, {"S": {"data": "A1:B1"}})
    assert plan.mode == "LIVE"
    assert plan.range_writes[0].values == [["x", "=A1+1"]]


def test_change_outside_region_defers():
    live = wb_bytes({"S": [[1, 10], [2, 20]]})
    prop = wb_bytes({"S": [[1, 10], [2, 99]]})
    # Region only covers row 1, but the change is in row 2.
    plan = wd.classify(prop, live, {"S": {"data": "A1:B1"}})
    assert plan.mode == "DEFER"
    assert "outside" in plan.reason


def test_no_region_map_defers_on_any_value_change():
    live = wb_bytes({"S": [[1]]})
    prop = wb_bytes({"S": [[2]]})
    plan = wd.classify(prop, live, None)
    assert plan.mode == "DEFER"


def test_empty_string_equals_none_no_change():
    live = wb_bytes({"S": [[1, None]]})
    prop = wb_bytes({"S": [[1, ""]]})
    plan = wd.classify(prop, live, {"S": {"data": "A1:B1"}})
    assert plan.mode == "LIVE"
    assert plan.reason == "already up to date"
    assert not plan.has_ops


# ---------------------------------------------------------------------------
# Region shrink
# ---------------------------------------------------------------------------


def test_region_shrink_clears_trailing_rows_via_nulls():
    live = wb_bytes({"S": [[1], [2], [3]]})
    prop = wb_bytes({"S": [[1]]})  # rows 2,3 now empty
    plan = wd.classify(prop, live, {"S": {"data": "A1:A3"}})
    assert plan.mode == "LIVE"
    w = plan.range_writes[0]
    assert w.address == "A1:A3"
    assert w.values == [[1], [None], [None]]


# ---------------------------------------------------------------------------
# Worksheet reconciliation
# ---------------------------------------------------------------------------


def test_reorder_is_live():
    live = wb_bytes({"A": [[1]], "B": [[2]]})
    prop = wb_bytes({"B": [[2]], "A": [[1]]})
    plan = wd.classify(prop, live, {"A": {"data": "A1:A1"}, "B": {"data": "A1:A1"}})
    assert plan.mode == "LIVE"
    assert [(o.name, o.position) for o in plan.reorders] == [("B", 0), ("A", 1)]
    assert not plan.range_writes


def test_add_plain_tab_is_live():
    live = wb_bytes({"A": [[1]]})
    prop = wb_bytes({"A": [[1]], "New": [["h"], [5]]})
    plan = wd.classify(prop, live, {"A": {"data": "A1:A1"}})
    assert plan.mode == "LIVE"
    assert len(plan.adds) == 1
    assert plan.adds[0].name == "New"
    assert plan.adds[0].writes[0].values == [["h"], [5]]


def test_delete_tab_is_live():
    live = wb_bytes({"A": [[1]], "Gone": [[9]]})
    prop = wb_bytes({"A": [[1]]})
    plan = wd.classify(prop, live, {"A": {"data": "A1:A1"}})
    assert plan.mode == "LIVE"
    assert [d.name for d in plan.deletes] == ["Gone"]


def test_rename_similar_content_is_live():
    body = [["alpha", 1], ["beta", 2], ["gamma", 3]]
    live = wb_bytes({"Old": body})
    prop = wb_bytes({"New": body})
    plan = wd.classify(prop, live, {"New": {"data": "A1:B3"}})
    assert plan.mode == "LIVE"
    assert [(r.old, r.new) for r in plan.renames] == [("Old", "New")]


def test_styled_new_tab_defers():
    def add_merge(wb):
        wb["New"].merge_cells("A1:B1")

    live = wb_bytes({"A": [[1]]})
    prop = wb_bytes({"A": [[1]], "New": [["title", None], [5, 6]]}, styler=add_merge)
    plan = wd.classify(prop, live, {"A": {"data": "A1:A1"}})
    assert plan.mode == "DEFER"
    assert "New" in plan.reason


def test_ambiguous_rename_defers():
    # One prop-only sheet equally similar to two live-only sheets -> ambiguous.
    shared = [["a"], ["b"], ["c"]]
    live = wb_bytes({"X": shared, "Y": shared})
    prop = wb_bytes({"Z": shared})
    plan = wd.classify(prop, live, {"Z": {"data": "A1:A3"}})
    assert plan.mode == "DEFER"
    assert "ambiguous" in plan.reason


# ---------------------------------------------------------------------------
# Format-only signal
# ---------------------------------------------------------------------------


def test_format_only_change_defers():
    live = wb_bytes({"S": [[1]]}, number_formats={"S": {(1, 1): "General"}})
    prop = wb_bytes({"S": [[1]]}, number_formats={"S": {(1, 1): "0.00"}})
    plan = wd.classify(prop, live, {"S": {"data": "A1:A1"}})
    assert plan.mode == "DEFER"
    assert "formatting" in plan.reason


def test_identical_workbooks_already_up_to_date():
    live = wb_bytes({"S": [[1, 2], [3, 4]]})
    prop = wb_bytes({"S": [[1, 2], [3, 4]]})
    plan = wd.classify(prop, live, {"S": {"data": "A1:B2"}})
    assert plan.mode == "LIVE"
    assert plan.reason == "already up to date"


def test_value_change_wins_over_format_difference():
    # A value changes in-region AND a number format differs -> still live-edit
    # the value (pragmatic: we leave formatting alone, don't defer).
    live = wb_bytes({"S": [[1]]}, number_formats={"S": {(1, 1): "General"}})
    prop = wb_bytes({"S": [[2]]}, number_formats={"S": {(1, 1): "0.00"}})
    plan = wd.classify(prop, live, {"S": {"data": "A1:A1"}})
    assert plan.mode == "LIVE"
    assert plan.range_writes[0].values == [[2]]


# ---------------------------------------------------------------------------
# Marshalling
# ---------------------------------------------------------------------------


def test_marshal_datetime_to_serial():
    out = wd.marshal_value(datetime(2020, 1, 1), wd.load_workbook(BytesIO(wb_bytes({"S": [[1]]}))).epoch)
    assert isinstance(out, (int, float)) and out > 40000


def test_marshal_bool_stays_bool():
    assert wd.marshal_value(True, None) is True
    assert wd.marshal_value(False, None) is False


def test_marshal_none_and_empty():
    assert wd.marshal_value(None, None) is None
    assert wd.marshal_value("", None) is None


def test_marshal_formula_passthrough():
    assert wd.marshal_value("=SUM(A1:A9)", None) == "=SUM(A1:A9)"


def test_marshal_nan_to_none():
    assert wd.marshal_value(float("nan"), None) is None
    assert wd.marshal_value(float("inf"), None) is None


def test_marshal_array_formula_defers():
    with pytest.raises(Exception):
        wd.marshal_value(ArrayFormula("A1:A2", "=B1:B2"), None)


def test_array_formula_in_proposed_defers_classify():
    def put_array(wb):
        wb["S"]["A1"] = ArrayFormula("A1", "=SUM(B1:B2)")

    live = wb_bytes({"S": [["x"]]})
    prop = wb_bytes({"S": [["x"]]}, styler=put_array)
    plan = wd.classify(prop, live, {"S": {"data": "A1:A1"}})
    assert plan.mode == "DEFER"
